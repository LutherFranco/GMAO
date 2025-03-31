import pandas as pd
import os
from collections import defaultdict
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# === 1. PARAMÈTRES ===
dossier = "./données/"
colonnes_ignorer = ['emplacement', 'asdu', 'nom poste reseau']
motifs_vides = ["", "VIDE"]

# === 2. Collecte des fichiers Excel ===
fichiers_excel = [f for f in os.listdir(dossier) if f.endswith(".xlsx") or f.endswith(".xlsm")]
liste_manquants = []

for fichier in fichiers_excel:
    equipement = os.path.splitext(fichier)[0].strip().replace(" ", "_")
    chemin = os.path.join(dossier, fichier)

    try:
        xl = pd.ExcelFile(chemin)
        if not xl.sheet_names:
            print(f"⛔ Aucun onglet dans {fichier} — ignoré")
            continue
        feuille = xl.sheet_names[0]
        df = xl.parse(feuille)
        df = df.loc[:, ~df.columns.str.contains("^unnamed", case=False)]
        print(f"📄 Lecture : {fichier} (onglet : {feuille})")
    except Exception as e:
        print(f"⛔ Erreur lecture fichier {fichier} : {e}")
        continue

    df.columns = df.columns.str.strip().str.replace('\n', ' ', regex=False).str.replace('\xa0', ' ', regex=False).str.lower()
    df.replace(motifs_vides, pd.NA, inplace=True)
    df["poste"] = df["emplacement"].astype(str).str[:5]

    colonnes_a_verifier = [col for col in df.columns if col not in colonnes_ignorer + ["poste"]]
    df_utiles = df[["poste", "emplacement"] + colonnes_a_verifier].copy()
    df_utiles.rename(columns={col: f"{equipement}_{col}" for col in colonnes_a_verifier}, inplace=True)

    for _, ligne in df_utiles.iterrows():
        poste = ligne["poste"]
        for col in ligne.index:
            if col not in ["emplacement", "poste"] and pd.isna(ligne[col]):
                liste_manquants.append({
                    "Poste": poste,
                    "Attribut manquant": col,
                    "Équipement": equipement
                })

# === 3. Données synthétiques
df_manquants = pd.DataFrame(liste_manquants)

df_synthese = df_manquants.groupby("Poste").agg({
    "Attribut manquant": lambda x: ", ".join(sorted(set(x))),
    "Équipement": lambda x: ", ".join(sorted(set(x)))
}).reset_index() if not df_manquants.empty else pd.DataFrame()

# === 4. Calcul de la complétude
compteur_total = defaultdict(int)
compteur_manquants = defaultdict(int)

for fichier in fichiers_excel:
    equipement = os.path.splitext(fichier)[0].strip().replace(" ", "_")
    chemin = os.path.join(dossier, fichier)

    try:
        xl = pd.ExcelFile(chemin)
        if not xl.sheet_names:
            continue
        df = xl.parse(xl.sheet_names[0])
        df = df.loc[:, ~df.columns.str.contains("^unnamed", case=False)]
    except:
        continue

    df.columns = df.columns.str.strip().str.replace('\n', ' ', regex=False).str.replace('\xa0', ' ', regex=False).str.lower()
    df.replace(motifs_vides, pd.NA, inplace=True)
    df["poste"] = df["emplacement"].astype(str).str[:5]
    colonnes_utiles = [col for col in df.columns if col not in colonnes_ignorer + ["poste"]]

    for _, ligne in df.iterrows():
        poste = ligne["poste"]
        compteur_total[poste] += len(colonnes_utiles)
        for col in colonnes_utiles:
            if pd.isna(ligne[col]):
                compteur_manquants[poste] += 1

def construire_ligne_completude(poste):
    total = compteur_total[poste]
    manquants = compteur_manquants.get(poste, 0)
    taux = round(100 * (1 - manquants / total), 1)
    niveau = (
        "🟢 Excellent" if taux >= 90 else
        "🟡 Moyen" if taux >= 70 else
        "🔴 Critique"
    )
    return {
        "Poste": poste,
        "Attributs attendus": total,
        "Attributs manquants": manquants,
        "Taux de complétude (%)": taux,
        "Niveau": niveau
    }

df_completude = pd.DataFrame([
    construire_ligne_completude(poste)
    for poste in sorted(compteur_total)
])

# === 5. Export Excel
with pd.ExcelWriter("résumé_attributs_manquants.xlsx", engine="openpyxl") as writer:
    if not df_manquants.empty:
        df_manquants.to_excel(writer, sheet_name="Détail", index=False)
        df_synthese.to_excel(writer, sheet_name="Synthèse par poste", index=False)
    df_completude.to_excel(writer, sheet_name="Complétude", index=False)

print("✅ Export terminé : résumé_attributs_manquants.xlsx")

# === 6. Mise en couleur de la complétude
wb = load_workbook("résumé_attributs_manquants.xlsx")
ws = wb["Complétude"]

# Trouver la colonne "Niveau"
col_niveau = None
for idx, cell in enumerate(ws[1], 1):
    if cell.value == "Niveau":
        col_niveau = idx
        break

# Couleurs
couleurs = {
    "🟢 Excellent": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # vert
    "🟡 Moyen": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),       # jaune
    "🔴 Critique": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # rouge
}

if col_niveau:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_niveau, max_col=col_niveau):
        cell = row[0]
        fill = couleurs.get(cell.value)
        if fill:
            cell.fill = fill

wb.save("résumé_attributs_manquants.xlsx")
print("🎨 Feuille Complétude mise en forme avec couleurs !")
